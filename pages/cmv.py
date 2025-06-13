import streamlit as st
import pandas as pd
import mysql.connector
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import calendar
import plotly.graph_objects as go
from decimal import Decimal


def mostrar_cmv():
    st.title("📊 Relatório CMV - Filiais")
    st.write("Esta é a página do relatório CMV por filial.")

    def conectar():
        return mysql.connector.connect(
            host='151.243.0.64',
            user='Usrflym_2265',
            password='FsyMdfYB74_p0',
            database='flymetrics'
        )

    hoje = datetime.now()
    col1, col2, col3 = st.columns(3)

    with col1:
        ano = st.selectbox("Ano", options=list(range(2024, hoje.year + 1)), index=list(range(2024, hoje.year + 1)).index(hoje.year))

    with col2:
        mes = st.selectbox("Mês", options=list(range(1, 13)), index=hoje.month - 1)

    with col3:
        filial = st.selectbox("Filial", ["RJ", "SP", "ES"])

    con = conectar()
    cursor = con.cursor()

    # Dias de faturamento no mês
    cursor.execute("SELECT COUNT(DISTINCT dia_ontem) FROM bf_cmv WHERE MONTH(dia_ontem) = %s AND YEAR(dia_ontem) = %s AND filial = %s", (mes, ano, filial))
    dias_faturamento = cursor.fetchone()[0]
    total_dias_mes = calendar.monthrange(ano, mes)[1]

    st.markdown(f"<div style='background-color:#264653;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>Dias com Faturamento</h5><h3 style='color:white;text-align:center;'>{dias_faturamento} de {total_dias_mes} dias</h3></div>", unsafe_allow_html=True)

    # Meta
    cursor.execute("SELECT meta FROM meta_rhino WHERE mes = %s AND ano = %s AND filial = %s", (mes, ano, filial))
    resultado = cursor.fetchone()
    meta = resultado[0] if resultado else None

    query_cmv = """
        SELECT 
          b.`dia_ontem` AS 'Data', 
          b.`faturamento` AS 'Faturamento', 
          b.qtd_pedidos AS 'Pedidos', 
          b.`markup` AS 'MKP', 
          b.`custo` AS 'CMV', 
          b.`compras` AS 'Compras',
          b.`qtd_devolucao` AS 'Qtd Devolvida', 
          b.`valor_devolucao` AS 'Devolução',
          b.`filial` AS 'Filial'
        FROM bf_cmv b
        WHERE MONTH(b.dia_ontem) = %s AND YEAR(b.dia_ontem) = %s AND b.filial = %s
        ORDER BY b.dia_ontem
    """

    df = pd.read_sql(query_cmv, con, params=(mes, ano, filial))
    valor_realizado = df["Faturamento"].sum()
    percentual = (valor_realizado / float(meta) * 100) if meta else 0

    st.subheader("🎯 Resultado vs Meta")
    if resultado:
        st.success("")
    else:
        st.warning("⚠️ Meta não cadastrada para este mês/ano/filial.")
        nova_meta = st.number_input("Informe a nova meta (R$)", min_value=0.01, step=100.0, format="%.2f")
        if st.button("Salvar nova meta"):
            if nova_meta and nova_meta > 0:
                cursor.execute(
                    "INSERT INTO meta_rhino (mes, ano, meta, filial) VALUES (%s, %s, %s, %s)",
                    (mes, ano, nova_meta, filial)
                )
                con.commit()
                st.success(f"✅ Meta de R$ {nova_meta:,.2f} salva com sucesso!")
            else:
                st.error("❌ Por favor, insira um valor válido para a meta.")
          

    with st.container():
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>META</h5><h5 style='color:white;text-align:center;'>R$ {meta if meta else 0:,.2f}</h3></div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>REALIZADO</h5><h5 style='color:white;text-align:center;'>R$ {valor_realizado:,.2f}</h3></div>", unsafe_allow_html=True)
        with col3:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>% ALCANÇADO</h5><h5 style='color:white;text-align:center;'>{percentual:.2f}%</h3></div>", unsafe_allow_html=True)

    st.subheader("📊 Relatório CMV")
    #st.dataframe(df)

    totais = df.drop(columns=["Data", "Filial"]).sum()
    totais_df = pd.DataFrame(totais).transpose()
    totais_df.insert(0, "Data", "TOTAL")
    totais_df.insert(8, "Filial", filial)
    st.dataframe(pd.concat([df, totais_df], ignore_index=True))

        # Resultado CMV (Custo - Compras)
    resultado_cmv = df["CMV"].sum() - df["Compras"].sum()
    seta = "⬆️" if resultado_cmv >= 0 else "⬇️"
    cor = "green" if resultado_cmv >= 0 else "red"

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"<div style='background-color:{cor};padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>Resultado CMV</h5><h5 style='color:white;text-align:center;'>{seta} R$ {resultado_cmv:,.2f}</h3></div>", unsafe_allow_html=True)
    with col2:
        # Estoque total
        cursor.execute("SELECT valor_estoque FROM bf_estoque_rhino WHERE MONTH(dia_ontem) = %s AND YEAR(dia_ontem) = %s AND filial = %s", (mes, ano, filial))
        resultado_estoque = cursor.fetchone()
        estoque_total = resultado_estoque[0] if resultado_estoque else 0
        st.markdown(f"<div style='background-color:#2a9d8f;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>Estoque Total</h5><h5 style='color:white;text-align:center;'>R$ {estoque_total:,.2f}</h3></div>", unsafe_allow_html=True)
    with col3:
        estoque_atualizado = estoque_total - Decimal(df["Compras"].sum())
        st.markdown(f"<div style='background-color:#e76f51;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>Estoque Atualizado</h5><h5 style='color:white;text-align:center;'>R$ {estoque_atualizado:,.2f}</h3></div>", unsafe_allow_html=True)

    # 🧮 Cálculos
    faturamento_total = df["Faturamento"].sum()
    pedidos_total = df["Pedidos"].sum()
    dias_com_faturamento = df[df["Faturamento"] > 0].shape[0]
    devolucao_total = df["Devolução"].sum()

    ticket_medio = faturamento_total / pedidos_total if pedidos_total else 0
    media_pedidos_dia = pedidos_total / dias_com_faturamento if dias_com_faturamento else 0
    porcentagem_devolucao = (devolucao_total / faturamento_total) * 100 if faturamento_total else 0

    # 🔳 Cartões com 4 colunas
    st.markdown("### 📈 Resultados de Venda")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(label="💰 Faturamento Total", value=f"R$ {faturamento_total:,.2f}")

    with col2:
        st.metric(label="🎟️ Ticket Médio", value=f"R$ {ticket_medio:,.2f}")

    with col3:
        st.metric(label="📦 Média de Pedidos por Dia", value=f"{media_pedidos_dia:.2f}")

    with col4:
        st.metric(label="🔁 % Devolução", value=f"{porcentagem_devolucao:.2f} %")


    # Comparação semanal de faturamento
    data_base = datetime(ano, mes, 1)
    semana_atual_fim = datetime(ano, mes, hoje.day)
    semana_atual_inicio = semana_atual_fim - timedelta(days=6)
    semana_anterior_fim = semana_atual_inicio - timedelta(days=1)
    semana_anterior_inicio = semana_anterior_fim - timedelta(days=6)

    cursor.execute("""
        SELECT SUM(faturamento) 
        FROM bf_cmv 
        WHERE dia_ontem BETWEEN %s AND %s AND filial = %s
    """, (semana_anterior_inicio, semana_anterior_fim, filial))
    resultado_anterior = cursor.fetchone()
    faturamento_semana_anterior = resultado_anterior[0] if resultado_anterior and resultado_anterior[0] else 0

    cursor.execute("""
        SELECT SUM(faturamento) 
        FROM bf_cmv 
        WHERE dia_ontem BETWEEN %s AND %s AND filial = %s
    """, (semana_atual_inicio, semana_atual_fim, filial))
    resultado_atual = cursor.fetchone()
    faturamento_semana_atual = resultado_atual[0] if resultado_atual and resultado_atual[0] else 0

    fig = go.Figure(data=[
        go.Bar(name='Semana Atual', x=['Faturamento'], y=[faturamento_semana_atual], marker_color='blue', text=[f"R$ {faturamento_semana_atual:,.2f}"], textposition='outside'),
        go.Bar(name='Semana Anterior', x=['Faturamento'], y=[faturamento_semana_anterior], marker_color='orange', text=[f"R$ {faturamento_semana_anterior:,.2f}"], textposition='outside')
    ])
    fig.update_layout(
        title='📊 Comparação Semanal de Faturamento',
        yaxis_title='R$',
        barmode='group',
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(color="white")
    )

    st.plotly_chart(fig, use_container_width=True)
    
        # Calcular projeção de faturamento
    faturamento_total = df["Faturamento"].sum()
    dias_com_faturamento = df[df["Faturamento"] > 0]["Data"].nunique()
    dias_totais_mes = pd.Period(f"{ano}-{mes:02}").days_in_month

    projecao_faturamento = (
        (faturamento_total / dias_com_faturamento) * dias_totais_mes
        if dias_com_faturamento > 0 else 0
    )

    # --- AQUI É ONDE VOCÊ PRECISA ATRIBUIR O VALOR DA META ---
    # Supondo que 'meta_valor' é o nome da sua variável que já contém a meta do banco
    #meta = meta_valor 

    # Verifique se 'meta' é um número e não é zero, antes de usar no cálculo
    # Isso é importante para evitar erros de divisão por zero ou type errors
    if not isinstance(meta, (int, float)):
        try:
            meta = float(meta) # Tenta converter para float
        except ValueError:
            st.error("Erro: O valor da meta não é um número válido. Verifique sua fonte de dados.")
            meta = 1 # Define um valor padrão para evitar erro, ou trate de outra forma
    elif meta == 0:
        st.warning("Aviso: A meta é zero. O percentual de projeção vs meta será 0%.")

    percentual_meta = (projecao_faturamento / float(meta) * 100) if meta else 0

    # Exibir cartões
    st.markdown("### 🔮 Projeção de Faturamento Mensal")
    col1, col2 = st.columns(2)

    with col1:
        st.metric("Faturamento Projetado", f"R$ {projecao_faturamento:,.2f}")

    with col2:
        st.metric("Projeção vs Meta", f"{percentual_meta:.2f}%")


    # Exportar Excel
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="CMV Rhino")
        ws = writer.sheets["CMV Rhino"]

        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

            

    buffer.seek(0)
    st.download_button("📥 Baixar Excel", buffer, file_name="cmv_rhino.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
if __name__ == "__main__":
    mostrar_cmv()