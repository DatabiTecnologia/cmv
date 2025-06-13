import streamlit as st
import pandas as pd
import mysql.connector
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

def mostrar_cmv():
    st.title("📊 Relatório CMV - Filiais")
    st.write("Esta é a página do relatório CMV por filial.")
    
    def conectar():
        return mysql.connector.connect(
            host='151.243.0.64',
            user='Usrflym_2265',
            password='FsyMdfYB74_p0',
            database='flymetrics'
        )

    hoje = datetime.now()
    col1, col2, col3 = st.columns(3)

    with col1:
        ano = st.selectbox("Ano", options=list(range(2024, hoje.year + 1)), index=list(range(2024, hoje.year + 1)).index(hoje.year))

    with col2:
        mes = st.selectbox("Mês", options=list(range(1, 13)), index=hoje.month - 1)

    with col3:
        filial = st.selectbox("Filial", ["RJ", "SP", "ES"])

    con = conectar()
    cursor = con.cursor()

    # Recupera meta
    cursor.execute("SELECT meta FROM meta_rhino WHERE mes = %s AND ano = %s AND filial = %s", (mes, ano, filial))
    resultado = cursor.fetchone()
    meta = resultado[0] if resultado else None

    # Consulta com JOIN incluindo o estoque
    query = """
        SELECT 
            c.dia_ontem AS 'Data',
            c.faturamento AS 'Faturamento',
            c.qtd_pedidos AS 'Pedidos',
            c.markup AS 'MKP',
            c.custo AS 'CMV',
            c.compras AS 'Compras',
            c.qtd_devolucao AS 'Qtd Devolvida',
            c.valor_devolucao AS 'Devolução',
            e.valor_estoque AS 'Estoque',
            c.filial AS 'Filial'
        FROM bf_cmv c
        LEFT JOIN bf_estoque_rhino e ON c.dia_ontem = e.dia_ontem AND c.filial = e.filial
        WHERE MONTH(c.dia_ontem) = %s AND YEAR(c.dia_ontem) = %s AND c.filial = %s
        ORDER BY c.dia_ontem
    """
    df = pd.read_sql(query, con, params=(mes, ano, filial))

    valor_realizado = df["Faturamento"].sum()
    percentual = (valor_realizado / float(meta) * 100) if meta else 0

    st.subheader("🎯 Resultado vs Meta")
    if resultado:
        st.success("")
    else:
        st.warning("⚠️ Meta não cadastrada para este mês/ano/filial.")
        nova_meta = st.number_input("Informe a nova meta (R$)", min_value=0.01, step=100.0, format="%.2f")
        if st.button("Salvar nova meta"):
            if nova_meta and nova_meta > 0:
                cursor.execute(
                    "INSERT INTO meta_rhino (mes, ano, meta, filial) VALUES (%s, %s, %s, %s)",
                    (mes, ano, nova_meta, filial)
                )
                con.commit()
                st.success(f"✅ Meta de R$ {nova_meta:,.2f} salva com sucesso!")
            else:
                st.error("❌ Por favor, insira um valor válido para a meta.")

    con.close()

    with st.container():
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>META</h5><h3 style='color:white;text-align:center;'>R$ {meta if meta else 0:,.2f}</h3></div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>REALIZADO</h5><h3 style='color:white;text-align:center;'>R$ {valor_realizado:,.2f}</h3></div>", unsafe_allow_html=True)
        with col3:
            st.markdown(f"<div style='background-color:black;padding:10px;border-radius:10px'><h5 style='color:white;text-align:center;'>% ALCANÇADO</h5><h3 style='color:white;text-align:center;'>{percentual:.2f}%</h3></div>", unsafe_allow_html=True)

    st.subheader("📊 Relatório CMV")
    st.dataframe(df)

    # Exportar Excel
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="CMV Rhino")
        ws = writer.sheets["CMV Rhino"]

        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer.seek(0)
    st.download_button("📥 Baixar Excel", buffer, file_name="cmv_rhino.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
